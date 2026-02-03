"""
Advanced Email Tool - Excel Tab
===============================
Tab for loading Excel files and mapping columns to email fields.
"""

from typing import Optional, Dict, List, Any
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QPushButton, QLabel, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QFileDialog,
    QMessageBox, QSplitter, QFrame, QSpinBox,
    QProgressDialog
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QColor, QBrush

from core.excel_handler import ExcelHandler
from core.validators import is_valid_email
from ui.components.dialogs import show_error, show_info, show_file_dialog, show_save_dialog
from ui.components.tab_navigation import TabNavigationBar
from ui.components.toggle_switch import LabeledToggleSwitch
from utils import get_logger

import config
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class TabExcel(QWidget):
    """
    Excel file loading and column mapping tab.
    
    Features:
    - Load Excel file
    - Select worksheet
    - Preview data
    - Map columns to email fields (To, CC, BCC, Identifier)
    """
    
    # Signals
    data_loaded = pyqtSignal(list, list)  # Emits (columns, data)
    mapping_changed = pyqtSignal(dict)  # Emits column mapping
    navigate_previous = pyqtSignal()  # Navigation signal
    navigate_next = pyqtSignal()  # Navigation signal
    
    def __init__(self, parent=None):
        """Initialize the Excel tab."""
        super().__init__(parent)
        
        self.logger = get_logger()
        self.excel_handler = ExcelHandler()
        self._file_path: Optional[str] = None
        self._columns: List[str] = []
        self._data: List[Dict[str, Any]] = []
        
        self._setup_ui()

    def _find_column(self, column_name: str) -> Optional[str]:
        """Find actual column name using case-insensitive match."""
        if not column_name:
            return None
        col_lower = column_name.lower()
        for col in self._columns:
            if col.lower() == col_lower:
                return col
        return None

    def _setup_ui(self) -> None:
        """Set up the user interface."""
        layout = QVBoxLayout(self)
        
        # File selection section
        file_group = QGroupBox("Excel File")
        file_layout = QVBoxLayout(file_group)
        
        # File path row
        path_layout = QHBoxLayout()
        
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("color: gray;")
        path_layout.addWidget(self.file_label, stretch=1)
        
        self.browse_btn = QPushButton("Browse...")
        self.browse_btn.clicked.connect(self._browse_file)
        path_layout.addWidget(self.browse_btn)
        
        self.reload_btn = QPushButton("Refresh")
        self.reload_btn.setToolTip("Reload file to pick up external changes")
        self.reload_btn.clicked.connect(self._reload_file)
        self.reload_btn.setEnabled(False)
        path_layout.addWidget(self.reload_btn)
        
        file_layout.addLayout(path_layout)
        
        # Sheet selection row
        sheet_layout = QHBoxLayout()
        
        sheet_label = QLabel("Worksheet:")
        sheet_layout.addWidget(sheet_label)
        
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(200)
        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        self.sheet_combo.setEnabled(False)
        sheet_layout.addWidget(self.sheet_combo)
        
        sheet_layout.addStretch()
        
        self.row_count_label = QLabel("")
        self.row_count_label.setStyleSheet("color: gray;")
        sheet_layout.addWidget(self.row_count_label)
        
        file_layout.addLayout(sheet_layout)
        
        # Data offset row (for data not starting at A1)
        offset_layout = QHBoxLayout()
        
        offset_label = QLabel("Data starts at:")
        offset_layout.addWidget(offset_label)
        
        offset_layout.addWidget(QLabel("Row:"))
        self.start_row_spin = QSpinBox()
        self.start_row_spin.setRange(1, 1000)
        self.start_row_spin.setValue(1)
        self.start_row_spin.setToolTip("Row number where headers are (1 = first row)")
        offset_layout.addWidget(self.start_row_spin)
        
        offset_layout.addSpacing(10)
        
        offset_layout.addWidget(QLabel("Column:"))
        self.start_col_spin = QSpinBox()
        self.start_col_spin.setRange(1, 100)
        self.start_col_spin.setValue(1)
        self.start_col_spin.setToolTip("Column number where data starts (1 = A, 2 = B, etc.)")
        offset_layout.addWidget(self.start_col_spin)
        
        offset_layout.addSpacing(20)
        
        self.apply_offset_btn = QPushButton("Apply")
        self.apply_offset_btn.clicked.connect(self._reload_with_offset)
        self.apply_offset_btn.setEnabled(False)
        offset_layout.addWidget(self.apply_offset_btn)
        
        offset_layout.addStretch()
        
        # Export template button
        self.export_template_btn = QPushButton("Export Blank Template")
        self.export_template_btn.clicked.connect(self._export_template)
        offset_layout.addWidget(self.export_template_btn)
        
        file_layout.addLayout(offset_layout)
        
        layout.addWidget(file_group)
        
        # Splitter for mapping and preview
        splitter = QSplitter(Qt.Horizontal)
        
        # Column mapping section
        mapping_group = QGroupBox("Column Mapping")
        mapping_layout = QVBoxLayout(mapping_group)
        
        mapping_info = QLabel(
            "Map Excel columns to email fields.\n"
            "Only 'To Email' is required."
        )
        mapping_info.setStyleSheet("color: gray; font-size: 10px;")
        mapping_layout.addWidget(mapping_info)
        
        # To Email (required)
        to_layout = QHBoxLayout()
        to_label = QLabel("To Email *:")
        to_label.setFixedWidth(100)
        to_label.setToolTip("Required. Column containing recipient email addresses.")
        to_layout.addWidget(to_label)
        self.to_combo = QComboBox()
        self.to_combo.setToolTip("Select the column with recipient email addresses")
        self.to_combo.currentIndexChanged.connect(self._on_mapping_changed)
        to_layout.addWidget(self.to_combo)
        mapping_layout.addLayout(to_layout)
        
        # CC (optional)
        cc_layout = QHBoxLayout()
        cc_label = QLabel("CC:")
        cc_label.setFixedWidth(100)
        cc_label.setToolTip("Optional. Carbon Copy recipients.")
        cc_layout.addWidget(cc_label)
        self.cc_combo = QComboBox()
        self.cc_combo.setToolTip("Column with CC email addresses (optional)")
        self.cc_combo.currentIndexChanged.connect(self._on_mapping_changed)
        cc_layout.addWidget(self.cc_combo)
        mapping_layout.addLayout(cc_layout)
        
        # BCC (optional)
        bcc_layout = QHBoxLayout()
        bcc_label = QLabel("BCC:")
        bcc_label.setFixedWidth(100)
        bcc_label.setToolTip("Optional. Blind Carbon Copy - recipients won't see each other.")
        bcc_layout.addWidget(bcc_label)
        self.bcc_combo = QComboBox()
        self.bcc_combo.setToolTip("Column with BCC email addresses (optional)")
        self.bcc_combo.currentIndexChanged.connect(self._on_mapping_changed)
        bcc_layout.addWidget(self.bcc_combo)
        mapping_layout.addLayout(bcc_layout)
        
        # Separator
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setStyleSheet("color: #ddd;")
        mapping_layout.addWidget(separator)
        
        # Identifier 1 (for attachment matching)
        id1_layout = QHBoxLayout()
        id1_label = QLabel("Identifier 1:")
        id1_label.setFixedWidth(100)
        id1_label.setToolTip(
            "Column used to match attachment files.\n"
            "Files with this value in their filename will be attached.\n"
            "e.g., If identifier is 'PAN001', file 'PAN001_report.pdf' matches."
        )
        id1_layout.addWidget(id1_label)
        self.identifier_combo = QComboBox()
        self.identifier_combo.setToolTip("Select column for attachment matching (e.g., PAN, Client Code)")
        self.identifier_combo.currentIndexChanged.connect(self._on_mapping_changed)
        id1_layout.addWidget(self.identifier_combo)
        mapping_layout.addLayout(id1_layout)
        
        # Identifier 2 (optional, for additional matching)
        id2_layout = QHBoxLayout()
        id2_label = QLabel("Identifier 2:")
        id2_label.setFixedWidth(100)
        id2_label.setToolTip("Optional second identifier for attachment matching.")
        id2_layout.addWidget(id2_label)
        self.identifier2_combo = QComboBox()
        self.identifier2_combo.setToolTip("Optional: Use two identifiers with AND/OR logic")
        self.identifier2_combo.currentIndexChanged.connect(self._on_mapping_changed)
        id2_layout.addWidget(self.identifier2_combo)
        mapping_layout.addLayout(id2_layout)
        
        # AND/OR logic selection
        logic_layout = QHBoxLayout()
        logic_label = QLabel("Match logic:")
        logic_label.setFixedWidth(100)
        logic_label.setToolTip("How to combine Identifier 1 and Identifier 2")
        logic_layout.addWidget(logic_label)

        self.logic_toggle = LabeledToggleSwitch("OR (Either)", "AND (Both)")
        self.logic_toggle.setToolTip(
            "OR = file matches either identifier\n"
            "AND = file must contain both identifiers"
        )
        self.logic_toggle.toggled.connect(self._on_mapping_changed)
        logic_layout.addWidget(self.logic_toggle)

        logic_layout.addStretch()
        mapping_layout.addLayout(logic_layout)
        
        mapping_layout.addStretch()
        
        # Validation button
        self.validate_btn = QPushButton("Validate Emails")
        self.validate_btn.clicked.connect(self._validate_emails)
        self.validate_btn.setEnabled(False)
        mapping_layout.addWidget(self.validate_btn)
        
        splitter.addWidget(mapping_group)
        
        # Data preview section
        preview_group = QGroupBox("Data Preview")
        preview_layout = QVBoxLayout(preview_group)
        
        # Validation status row
        validation_row = QHBoxLayout()
        self.validation_label = QLabel("")
        validation_row.addWidget(self.validation_label)
        validation_row.addStretch()
        
        self.preview_count_label = QLabel("")
        self.preview_count_label.setStyleSheet("color: gray;")
        validation_row.addWidget(self.preview_count_label)
        preview_layout.addLayout(validation_row)
        
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.setToolTip("Red rows indicate invalid email addresses")
        preview_layout.addWidget(self.preview_table)
        
        splitter.addWidget(preview_group)
        
        # Set splitter proportions
        splitter.setSizes([300, 500])

        layout.addWidget(splitter, stretch=1)

        # Navigation bar (first tab - no Previous button)
        self.nav_bar = TabNavigationBar(show_previous=False, show_next=True)
        self.nav_bar.next_clicked.connect(self.navigate_next.emit)
        layout.addWidget(self.nav_bar)

    def _browse_file(self) -> None:
        """Open file dialog to select Excel file."""
        file_path = show_file_dialog(
            self,
            title="Select Excel File",
            filter=config.EXCEL_FILTER
        )
        
        if file_path:
            self._load_file(file_path)
    
    def _load_file(self, file_path: str) -> None:
        """
        Load an Excel file.
        
        Args:
            file_path: Path to Excel file
        """
        self.logger.info(f"Loading file: {file_path}")
        
        success, error = self.excel_handler.load_file(file_path)
        
        if not success:
            show_error(self, "Error Loading File", error)
            return
        
        self._file_path = file_path
        self._update_ui_after_load()
    
    def _update_ui_after_load(self) -> None:
        """Update UI after successful file load."""
        # Update file label
        import os
        filename = os.path.basename(self._file_path)
        self.file_label.setText(filename)
        self.file_label.setToolTip(self._file_path)
        self.file_label.setStyleSheet("")
        
        # Update sheet combo
        sheets = self.excel_handler.get_sheet_names()
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(sheets)
        self.sheet_combo.blockSignals(False)
        self.sheet_combo.setEnabled(True)
        
        # Enable buttons
        self.reload_btn.setEnabled(True)
        self.validate_btn.setEnabled(True)
        self.apply_offset_btn.setEnabled(True)
        
        # Load data from first sheet
        self._load_sheet_data()
    
    def _on_sheet_changed(self, index: int) -> None:
        """Handle worksheet selection change."""
        if index < 0 or not self._file_path:
            return
        
        sheet_name = self.sheet_combo.currentText()
        success, error = self.excel_handler.load_file(self._file_path, sheet_name)
        
        if success:
            self._load_sheet_data()
        else:
            show_error(self, "Error Loading Sheet", error)
    
    def _load_sheet_data(self) -> None:
        """Load data from current sheet and update UI."""
        self._columns = self.excel_handler.get_columns()
        self._data = self.excel_handler.get_data(filtered=False)
        
        # Update row count
        count = len(self._data)
        self.row_count_label.setText(f"{count} row{'s' if count != 1 else ''}")
        
        # Update column mappings
        self._populate_column_combos()
        
        # Update preview table
        self._populate_preview_table()
        
        # Emit signal
        self.data_loaded.emit(self._columns, self._data)
    
    def _populate_column_combos(self) -> None:
        """Populate column mapping dropdowns."""
        # Block signals during population
        combos = [self.to_combo, self.cc_combo, self.bcc_combo, self.identifier_combo, self.identifier2_combo]
        for combo in combos:
            combo.blockSignals(True)
            combo.clear()
        
        # Add empty option for optional fields
        self.cc_combo.addItem("-- None --")
        self.bcc_combo.addItem("-- None --")
        self.identifier_combo.addItem("-- None --")
        self.identifier2_combo.addItem("-- None --")
        
        # Add columns
        for col in self._columns:
            for combo in combos:
                combo.addItem(col)
        
        # Auto-detect email column
        self._auto_detect_columns()
        
        # Unblock signals
        for combo in combos:
            combo.blockSignals(False)
        
        # Trigger mapping changed
        self._on_mapping_changed()
    
    def _auto_detect_columns(self) -> None:
        """Try to auto-detect column mappings based on names."""
        columns_lower = {col.lower(): col for col in self._columns}
        
        # Detect email column
        email_patterns = ['email', 'e-mail', 'mail', 'to', 'recipient']
        for pattern in email_patterns:
            for col_lower, col in columns_lower.items():
                if pattern in col_lower:
                    idx = self._columns.index(col)
                    self.to_combo.setCurrentIndex(idx)
                    break
            else:
                continue
            break
        
        # Detect identifier column
        id_patterns = ['identifier', 'id', 'code', 'pan', 'gstin', 'client']
        for pattern in id_patterns:
            for col_lower, col in columns_lower.items():
                if pattern in col_lower:
                    idx = self._columns.index(col) + 1  # +1 for "-- None --"
                    self.identifier_combo.setCurrentIndex(idx)
                    break
            else:
                continue
            break
    
    def _populate_preview_table(self, max_rows: int = 100) -> None:
        """
        Populate the preview table with data.
        Validates emails and highlights invalid ones in red.
        
        Args:
            max_rows: Maximum rows to show in preview
        """
        self.preview_table.clear()
        
        if not self._data:
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            self._update_validation_status(0, 0)
            return
        
        # Set up columns
        self.preview_table.setColumnCount(len(self._columns))
        self.preview_table.setHorizontalHeaderLabels(self._columns)
        
        # Limit rows for preview
        preview_data = self._data[:max_rows]
        self.preview_table.setRowCount(len(preview_data))
        
        # Get email column for validation (case-insensitive)
        email_column = self.to_combo.currentText() if hasattr(self, 'to_combo') else None
        actual_email_col = self._find_column(email_column)
        email_col_idx = self._columns.index(actual_email_col) if actual_email_col else -1
        
        # Get theme-aware colors for validation
        try:
            from utils.theme_manager import get_theme_manager
            theme_manager = get_theme_manager()
            is_dark = theme_manager.is_dark_mode() if theme_manager else False
        except:
            is_dark = False
        
        if is_dark:
            invalid_bg = QBrush(QColor("#4A1F1F"))  # Dark red
            invalid_fg = QColor("#F48771")  # Light red text
        else:
            invalid_bg = QBrush(QColor("#FFEBEE"))  # Light red
            invalid_fg = QColor("#DC3545")  # Red text
        
        # Track invalid count
        invalid_count = 0
        
        # Populate cells
        for row_idx, row_data in enumerate(preview_data):
            for col_idx, col_name in enumerate(self._columns):
                value = str(row_data.get(col_name, ""))
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                
                # Validate email column
                if col_idx == email_col_idx and value.strip():
                    if not is_valid_email(value.strip()):
                        item.setBackground(invalid_bg)
                        item.setForeground(invalid_fg)
                        item.setToolTip("Invalid email format")
                        invalid_count += 1
                
                self.preview_table.setItem(row_idx, col_idx, item)
        
        # Resize columns to content
        self.preview_table.resizeColumnsToContents()
        
        # Update validation status
        self._update_validation_status(len(preview_data), invalid_count)
    
    def _update_validation_status(self, total: int, invalid: int) -> None:
        """Update the validation status label."""
        if not hasattr(self, 'validation_label'):
            return
        
        # Get theme-aware colors
        try:
            from utils.theme_manager import get_theme_manager
            theme_manager = get_theme_manager()
            is_dark = theme_manager.is_dark_mode() if theme_manager else False
        except:
            is_dark = False
        
        success_color = "#4EC9B0" if is_dark else "#28A745"
        error_color = "#F48771" if is_dark else "#DC3545"
        
        if total == 0:
            self.validation_label.setText("")
            self.validation_label.setStyleSheet("")
        elif invalid == 0:
            self.validation_label.setText(f"✓ All {total} emails valid")
            self.validation_label.setStyleSheet(f"color: {success_color}; font-weight: bold;")
        else:
            self.validation_label.setText(f"⚠ {invalid} invalid email(s)")
            self.validation_label.setStyleSheet(f"color: {error_color}; font-weight: bold;")
    
    def _on_mapping_changed(self) -> None:
        """Handle column mapping changes."""
        mapping = self.get_column_mapping()
        self.mapping_changed.emit(mapping)
    
    def _validate_emails(self) -> None:
        """Validate email addresses in the mapped column."""
        email_column = self.to_combo.currentText()
        if not email_column:
            show_error(self, "Validation Error", "Please select an email column first.")
            return
        
        issues_idx, issues_msg = self.excel_handler.validate_email_column(email_column)
        
        if not issues_msg:
            show_info(
                self,
                "Validation Complete",
                f"All {len(self._data)} email addresses are valid!"
            )
        else:
            # Show issues (limit to first 20)
            display_issues = issues_msg[:20]
            if len(issues_msg) > 20:
                display_issues.append(f"... and {len(issues_msg) - 20} more issues")
            
            QMessageBox.warning(
                self,
                "Email Validation Issues",
                f"Found {len(issues_msg)} issue(s):\n\n" + "\n".join(display_issues)
            )
    
    def _reload_file(self) -> None:
        """Reload the current file to pick up external changes."""
        if not self._file_path:
            return

        # Store current state
        current_sheet = self.sheet_combo.currentText()
        current_mapping = self.get_column_mapping()
        current_row_offset = self.start_row_spin.value()
        current_col_offset = self.start_col_spin.value()

        # Reload file with current offsets
        if current_row_offset > 1 or current_col_offset > 1:
            success, error = self.excel_handler.load_file_with_offset(
                self._file_path,
                current_sheet,
                start_row=current_row_offset,
                start_col=current_col_offset
            )
        else:
            success, error = self.excel_handler.load_file(self._file_path, current_sheet)

        if success:
            self._load_sheet_data()
            # Restore column mapping if columns still exist
            self._restore_column_mapping(current_mapping)
            show_info(self, "Refreshed", "File reloaded with latest changes.")
        else:
            show_error(self, "Refresh Error", error)

    def _restore_column_mapping(self, mapping: Dict[str, Optional[str]]) -> None:
        """
        Restore column mapping after reload if columns still exist.

        Args:
            mapping: Previous column mapping to restore
        """
        def restore_combo(combo: QComboBox, value: Optional[str], has_none: bool = False):
            if not value:
                if has_none:
                    combo.setCurrentIndex(0)
                return
            actual_col = self._find_column(value)
            if actual_col:
                idx = combo.findText(actual_col)
                if idx >= 0:
                    combo.setCurrentIndex(idx)

        # Block signals during restoration
        combos = [self.to_combo, self.cc_combo, self.bcc_combo, self.identifier_combo, self.identifier2_combo]
        for combo in combos:
            combo.blockSignals(True)

        restore_combo(self.to_combo, mapping.get('to'))
        restore_combo(self.cc_combo, mapping.get('cc'), has_none=True)
        restore_combo(self.bcc_combo, mapping.get('bcc'), has_none=True)
        restore_combo(self.identifier_combo, mapping.get('identifier'), has_none=True)
        restore_combo(self.identifier2_combo, mapping.get('identifier2'), has_none=True)

        # Restore logic toggle (True = AND, False = OR)
        self.logic_toggle.setCheckedNoSignal(mapping.get('identifier_logic') == 'AND')

        # Unblock signals
        for combo in combos:
            combo.blockSignals(False)

        # Trigger mapping changed
        self._on_mapping_changed()
    
    def get_column_mapping(self) -> Dict[str, Optional[str]]:
        """
        Get the current column mapping.
        
        Returns:
            Dictionary with mapping for each field
        """
        def get_value(combo: QComboBox) -> Optional[str]:
            text = combo.currentText()
            if text and text != "-- None --":
                return text
            return None
        
        return {
            'to': get_value(self.to_combo),
            'cc': get_value(self.cc_combo),
            'bcc': get_value(self.bcc_combo),
            'identifier': get_value(self.identifier_combo),
            'identifier2': get_value(self.identifier2_combo),
            'identifier_logic': 'AND' if self.logic_toggle.isChecked() else 'OR',
        }
    
    def set_column_mapping(self, mapping: Dict[str, Optional[str]]) -> None:
        """
        Set the column mapping.
        
        Args:
            mapping: Dictionary with column mappings
        """
        def set_combo(combo: QComboBox, value: Optional[str], has_none: bool = False):
            actual_col = self._find_column(value) if value else None
            if actual_col:
                idx = combo.findText(actual_col)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
            elif has_none:
                combo.setCurrentIndex(0)  # "-- None --"
        
        set_combo(self.to_combo, mapping.get('to'))
        set_combo(self.cc_combo, mapping.get('cc'), has_none=True)
        set_combo(self.bcc_combo, mapping.get('bcc'), has_none=True)
        set_combo(self.identifier_combo, mapping.get('identifier'), has_none=True)
    
    def get_columns(self) -> List[str]:
        """
        Get list of column names.
        
        Returns:
            List of column names
        """
        return self._columns.copy()
    
    def get_data(self) -> List[Dict[str, Any]]:
        """
        Get the loaded data.
        
        Returns:
            List of row dictionaries
        """
        return self._data.copy()
    
    def get_excel_handler(self) -> ExcelHandler:
        """
        Get the Excel handler instance.
        
        Returns:
            ExcelHandler instance
        """
        return self.excel_handler
    
    def get_file_path(self) -> Optional[str]:
        """
        Get the loaded file path.
        
        Returns:
            File path or None
        """
        return self._file_path
    
    def is_data_loaded(self) -> bool:
        """
        Check if data is loaded.
        
        Returns:
            True if data is loaded
        """
        return len(self._data) > 0
    
    def _reload_with_offset(self) -> None:
        """Reload current file with specified row/column offset."""
        if not self._file_path:
            show_error(self, "Error", "No file loaded.")
            return
        
        start_row = self.start_row_spin.value()
        start_col = self.start_col_spin.value()
        
        # Load with offset
        sheet_name = self.sheet_combo.currentText() if self.sheet_combo.currentIndex() >= 0 else None
        success, error = self.excel_handler.load_file_with_offset(
            self._file_path, 
            sheet_name,
            start_row=start_row,
            start_col=start_col
        )
        
        if success:
            self._load_sheet_data()
            show_info(self, "Reloaded", f"Data reloaded starting from row {start_row}, column {start_col}.")
        else:
            show_error(self, "Error", error)
    
    def _export_template(self) -> None:
        """Export a blank Excel template with sample structure."""
        save_path = show_save_dialog(
            self,
            "Save Excel Template",
            "Excel Files (*.xlsx)",
            default_name="email_template.xlsx"
        )
        
        if not save_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recipients"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            # Headers
            headers = ["Name", "Email", "CC", "Identifier", "Custom1", "Custom2"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            # Sample data
            sample_data = [
                ["John Doe", "john@example.com", "", "PAN001", "Value1", "Value2"],
                ["Jane Smith", "jane@example.com", "manager@example.com", "PAN002", "Value1", "Value2"],
                ["", "", "", "", "", ""],
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Adjust column widths
            column_widths = [20, 30, 30, 15, 15, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Add instructions sheet
            ws_help = wb.create_sheet(title="Instructions")
            instructions = [
                ["Advanced Email Tool - Template Instructions"],
                [""],
                ["Required Columns:"],
                ["- Email: Recipient email address (required)"],
                [""],
                ["Optional Columns:"],
                ["- Name: Recipient name (use in email as {Name})"],
                ["- CC: CC email address"],
                ["- Identifier: For matching attachments (e.g., PAN, Client Code)"],
                ["- Custom columns: Any additional data to use as {ColumnName} in templates"],
                [""],
                ["Tips:"],
                ["- First row must be headers"],
                ["- Multiple emails in one cell: separate with semicolon (;)"],
                ["- Variables in email template use format: {ColumnName}"],
                ["- Identifier is matched as substring in filenames"],
            ]
            
            for row_idx, row in enumerate(instructions, 1):
                cell = ws_help.cell(row=row_idx, column=1, value=row[0] if row else "")
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row and row[0].endswith(":"):
                    cell.font = Font(bold=True)
            
            ws_help.column_dimensions["A"].width = 70
            
            # Save
            wb.save(save_path)
            wb.close()
            
            show_info(self, "Template Exported", f"Template saved to:\n{save_path}")
            self.logger.info(f"Template exported to: {save_path}")
            
        except Exception as e:
            show_error(self, "Export Error", f"Could not export template: {e}")
            self.logger.error(f"Template export error: {e}")
    
    def load_file_path(self, file_path: str) -> bool:
        """
        Load a specific file path (for session restore).
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            True if loaded successfully
        """
        import os
        if os.path.exists(file_path):
            self._load_file(file_path)
            return self.is_data_loaded()
        return False
