"""
Advanced Email Tool - Excel Handler
===================================
Handles loading, validating, filtering, and sorting Excel data.
"""

from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from utils import get_logger


class ExcelHandler:
    """
    Handles all Excel file operations.
    
    Loads Excel files and provides methods for filtering,
    sorting, and accessing data.
    """
    
    def __init__(self):
        """Initialize the Excel handler."""
        self.logger = get_logger()
        self.workbook = None
        self.worksheet = None
        self.file_path: Optional[str] = None
        self.columns: List[str] = []
        self.data: List[Dict[str, Any]] = []
        self.filtered_data: List[Dict[str, Any]] = []
        self.row_count: int = 0
    
    def load_file(self, file_path: str, sheet_name: str = None) -> Tuple[bool, str]:
        """
        Load an Excel file.
        
        Args:
            file_path: Path to Excel file (.xlsx)
            sheet_name: Name of sheet to load (default: first sheet)
            
        Returns:
            Tuple of (success, error_message)
        """
        try:
            self.logger.info(f"Loading Excel file: {file_path}")
            
            # Validate file exists
            path = Path(file_path)
            if not path.exists():
                return False, f"File not found: {file_path}"
            
            if not path.suffix.lower() in ['.xlsx', '.xlsm']:
                return False, "File must be .xlsx or .xlsm format"
            
            # Load workbook
            self.workbook = openpyxl.load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True  # Get values, not formulas
            )
            
            # Select worksheet
            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    return False, f"Sheet '{sheet_name}' not found"
                self.worksheet = self.workbook[sheet_name]
            else:
                self.worksheet = self.workbook.active
            
            # Read data
            self._read_data()
            
            self.file_path = file_path
            self.logger.info(f"Loaded {self.row_count} rows with {len(self.columns)} columns")
            
            return True, ""
            
        except openpyxl.utils.exceptions.InvalidFileException:
            return False, "Invalid Excel file format"
        except PermissionError:
            return False, "Permission denied - file may be open in another program"
        except Exception as e:
            self.logger.error(f"Error loading Excel: {e}", exc_info=True)
            return False, f"Error loading file: {e}"
    
    def _read_data(self, start_row: int = 1, start_col: int = 1) -> None:
        """
        Read data from the active worksheet.
        
        Args:
            start_row: Row number where headers are (1-based)
            start_col: Column number where data starts (1-based)
        """
        self.columns = []
        self.data = []
        
        rows = list(self.worksheet.iter_rows(values_only=True))
        
        if not rows:
            self.row_count = 0
            return
        
        # Adjust for 0-based indexing
        header_row_idx = start_row - 1
        col_offset = start_col - 1
        
        if header_row_idx >= len(rows):
            self.row_count = 0
            return
        
        # Get headers from specified row
        header_row = rows[header_row_idx]
        self.columns = []
        for i, cell in enumerate(header_row):
            if i >= col_offset:
                col_name = str(cell) if cell else f"Column_{i - col_offset + 1}"
                self.columns.append(col_name)
        
        # Remaining rows are data (after header row)
        for row in rows[header_row_idx + 1:]:
            # Get only columns from offset onwards
            row_values = row[col_offset:] if col_offset < len(row) else []
            
            if all(cell is None for cell in row_values):
                continue  # Skip empty rows
            
            row_dict = {}
            for i, value in enumerate(row_values):
                if i < len(self.columns):
                    col_name = self.columns[i]
                    row_dict[col_name] = value if value is not None else ""
            
            self.data.append(row_dict)
        
        self.row_count = len(self.data)
        self.filtered_data = self.data.copy()
    
    def load_file_with_offset(
        self, 
        file_path: str, 
        sheet_name: str = None,
        start_row: int = 1,
        start_col: int = 1
    ) -> Tuple[bool, str]:
        """
        Load an Excel file with data starting from specified row/column.
        
        Args:
            file_path: Path to Excel file (.xlsx)
            sheet_name: Name of sheet to load (default: first sheet)
            start_row: Row number where headers are (1-based)
            start_col: Column number where data starts (1-based)
            
        Returns:
            Tuple of (success, error_message)
        """
        try:
            self.logger.info(f"Loading Excel file with offset: {file_path} (row={start_row}, col={start_col})")
            
            path = Path(file_path)
            if not path.exists():
                return False, f"File not found: {file_path}"
            
            if not path.suffix.lower() in ['.xlsx', '.xlsm']:
                return False, "File must be .xlsx or .xlsm format"
            
            self.workbook = openpyxl.load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True
            )
            
            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    return False, f"Sheet '{sheet_name}' not found"
                self.worksheet = self.workbook[sheet_name]
            else:
                self.worksheet = self.workbook.active
            
            self._read_data(start_row, start_col)
            
            self.file_path = file_path
            self.logger.info(f"Loaded {self.row_count} rows with {len(self.columns)} columns")
            
            return True, ""
            
        except openpyxl.utils.exceptions.InvalidFileException:
            return False, "Invalid Excel file format"
        except PermissionError:
            return False, "Permission denied - file may be open in another program"
        except Exception as e:
            self.logger.error(f"Error loading Excel: {e}", exc_info=True)
            return False, f"Error loading file: {e}"
    
    def get_columns(self) -> List[str]:
        """
        Get list of column names.
        
        Returns:
            List of column name strings
        """
        return self.columns.copy()
    
    def get_data(self, filtered: bool = True) -> List[Dict[str, Any]]:
        """
        Get data rows.
        
        Args:
            filtered: If True, return filtered data; else return all data
            
        Returns:
            List of row dictionaries
        """
        if filtered:
            return self.filtered_data.copy()
        return self.data.copy()
    
    def get_row(self, index: int, filtered: bool = True) -> Optional[Dict[str, Any]]:
        """
        Get a single row by index.
        
        Args:
            index: Row index (0-based)
            filtered: Whether to use filtered data
            
        Returns:
            Row dictionary or None if index out of range
        """
        data = self.filtered_data if filtered else self.data
        if 0 <= index < len(data):
            return data[index].copy()
        return None
    
    def get_column_values(self, column_name: str, filtered: bool = True) -> List[Any]:
        """
        Get all values for a specific column.
        
        Args:
            column_name: Name of the column
            filtered: Whether to use filtered data
            
        Returns:
            List of values
        """
        data = self.filtered_data if filtered else self.data
        return [row.get(column_name, "") for row in data]
    
    def get_unique_values(self, column_name: str, filtered: bool = True) -> List[Any]:
        """
        Get unique values for a column.
        
        Args:
            column_name: Name of the column
            filtered: Whether to use filtered data
            
        Returns:
            List of unique values (preserves order)
        """
        values = self.get_column_values(column_name, filtered)
        seen = set()
        unique = []
        for v in values:
            v_str = str(v) if v else ""
            if v_str not in seen:
                seen.add(v_str)
                unique.append(v)
        return unique
    
    def filter_data(self, column_name: str, value: Any, exact_match: bool = True) -> int:
        """
        Filter data by column value.
        
        Args:
            column_name: Column to filter on
            value: Value to filter for
            exact_match: If True, exact match; else contains match
            
        Returns:
            Number of rows after filtering
        """
        if column_name not in self.columns:
            self.logger.warning(f"Filter column '{column_name}' not found")
            return len(self.filtered_data)
        
        value_str = str(value).lower() if value else ""
        
        if exact_match:
            self.filtered_data = [
                row for row in self.data
                if str(row.get(column_name, "")).lower() == value_str
            ]
        else:
            self.filtered_data = [
                row for row in self.data
                if value_str in str(row.get(column_name, "")).lower()
            ]
        
        self.logger.debug(f"Filtered by {column_name}='{value}': {len(self.filtered_data)} rows")
        return len(self.filtered_data)
    
    def clear_filter(self) -> int:
        """
        Clear all filters and restore original data.
        
        Returns:
            Number of rows (total)
        """
        self.filtered_data = self.data.copy()
        return len(self.filtered_data)
    
    def sort_data(self, column_name: str, ascending: bool = True) -> None:
        """
        Sort filtered data by column.
        
        Args:
            column_name: Column to sort by
            ascending: Sort direction
        """
        if column_name not in self.columns:
            self.logger.warning(f"Sort column '{column_name}' not found")
            return
        
        def sort_key(row):
            val = row.get(column_name, "")
            if val is None:
                return ""
            return str(val).lower()
        
        self.filtered_data.sort(key=sort_key, reverse=not ascending)
    
    def get_row_count(self, filtered: bool = True) -> int:
        """
        Get number of rows.
        
        Args:
            filtered: Whether to count filtered data
            
        Returns:
            Row count
        """
        if filtered:
            return len(self.filtered_data)
        return len(self.data)
    
    def get_sheet_names(self) -> List[str]:
        """
        Get list of sheet names in workbook.
        
        Returns:
            List of sheet names
        """
        if self.workbook:
            return self.workbook.sheetnames
        return []
    
    def validate_email_column(self, column_name: str) -> Tuple[List[int], List[str]]:
        """
        Validate emails in a column and find issues.
        
        Args:
            column_name: Name of column containing emails
            
        Returns:
            Tuple of (row_indices_with_issues, error_messages)
        """
        from core.validators import validate_email_list, suggest_email_correction
        
        issues_indices = []
        issues_messages = []
        
        for idx, row in enumerate(self.filtered_data):
            email_value = str(row.get(column_name, ""))
            
            if not email_value:
                issues_indices.append(idx)
                issues_messages.append(f"Row {idx + 2}: Empty email")
                continue
            
            is_valid, valid_emails, errors = validate_email_list(email_value)
            
            if not is_valid:
                issues_indices.append(idx)
                issues_messages.append(f"Row {idx + 2}: {'; '.join(errors)}")
            else:
                # Check for typo suggestions
                for email in valid_emails:
                    suggestion = suggest_email_correction(email)
                    if suggestion:
                        issues_messages.append(
                            f"Row {idx + 2}: Did you mean {suggestion}? (currently: {email})"
                        )
        
        return issues_indices, issues_messages
    
    def close(self) -> None:
        """Close the workbook and release resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
        
        self.file_path = None
        self.columns = []
        self.data = []
        self.filtered_data = []
        self.row_count = 0
