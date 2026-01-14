"""
Advanced Email Tool - Attachment Matcher
========================================
Matches files to clients based on identifier patterns.
Uses exact substring matching in filenames (case-insensitive).
"""

import os
from typing import List, Dict, Tuple, Optional, Any
from pathlib import Path
from datetime import datetime

from utils import get_logger, scan_directory, check_path_accessible, format_bytes

import config


class AttachmentMatcher:
    """
    Matches files to clients based on identifiers.
    
    Scans a directory once (upfront indexing) and then
    matches files to identifiers using exact substring matching.
    """
    
    def __init__(self):
        """Initialize the attachment matcher."""
        self.logger = get_logger()
        self.directory: Optional[str] = None
        self.recursive: bool = False
        self.file_index: List[Tuple[str, str, int]] = []  # (full_path, filename, size)
        self.matches: Dict[str, List[Tuple[str, str, int]]] = {}  # identifier -> files
        self.scan_time: Optional[datetime] = None
        self.total_files: int = 0
        self.total_size: int = 0
    
    def set_directory(
        self, 
        directory: str, 
        recursive: bool = False
    ) -> Tuple[bool, str]:
        """
        Set the directory to scan for attachments.
        
        Args:
            directory: Path to attachment directory
            recursive: Whether to scan subdirectories
            
        Returns:
            Tuple of (success, error_message)
        """
        # Validate directory
        is_accessible, error = check_path_accessible(directory)
        if not is_accessible:
            return False, f"Cannot access directory: {error}"
        
        self.directory = directory
        self.recursive = recursive
        self.file_index = []
        self.matches = {}
        self.scan_time = None
        
        self.logger.info(f"Attachment directory set: {directory} (recursive={recursive})")
        return True, ""
    
    def scan(self, extensions: List[str] = None) -> Tuple[int, str]:
        """
        Scan the directory and build file index.
        
        Args:
            extensions: File extensions to include (None = all)
            
        Returns:
            Tuple of (file_count, status_message)
        """
        if not self.directory:
            return 0, "No directory set"
        
        self.logger.info(f"Scanning directory: {self.directory}")
        
        self.file_index = []
        self.total_size = 0
        
        try:
            for full_path, filename, size in scan_directory(
                self.directory, 
                self.recursive,
                extensions
            ):
                self.file_index.append((full_path, filename, size))
                self.total_size += size
            
            self.total_files = len(self.file_index)
            self.scan_time = datetime.now()
            
            formatted_size = format_bytes(self.total_size)
            status = f"Found {self.total_files} files ({formatted_size})"
            
            self.logger.info(status)
            return self.total_files, status
            
        except Exception as e:
            self.logger.error(f"Error scanning directory: {e}", exc_info=True)
            return 0, f"Error scanning: {e}"
    
    def match_identifier(self, identifier: str) -> List[Tuple[str, str, int]]:
        """
        Find all files matching an identifier.
        
        Uses exact substring matching (case-insensitive) on filenames.
        
        Args:
            identifier: Identifier string to match
            
        Returns:
            List of tuples (full_path, filename, size_bytes)
        """
        if not identifier:
            return []
        
        identifier = str(identifier).strip()
        
        # Check cache first
        if identifier in self.matches:
            return self.matches[identifier]
        
        # Find matches
        matched = []
        identifier_lower = identifier.lower()
        for full_path, filename, size in self.file_index:
            # Case-insensitive substring match in filename
            if identifier_lower in filename.lower():
                matched.append((full_path, filename, size))
        
        # Cache result
        self.matches[identifier] = matched
        
        self.logger.debug(f"Identifier '{identifier}' matched {len(matched)} file(s)")
        
        return matched
    
    def match_all_identifiers(
        self, 
        identifiers: List[str]
    ) -> Dict[str, List[Tuple[str, str, int]]]:
        """
        Match all identifiers at once.
        
        Args:
            identifiers: List of identifier strings
            
        Returns:
            Dictionary mapping identifiers to matched files
        """
        results = {}
        for identifier in identifiers:
            if identifier:
                results[str(identifier).strip()] = self.match_identifier(identifier)
        return results
    
    def match_multiple_identifiers(
        self, 
        identifiers: List[str]
    ) -> List[Tuple[str, str, int]]:
        """
        Find files matching ANY of the given identifiers (OR logic).
        
        Args:
            identifiers: List of identifiers to search for
            
        Returns:
            List of unique matched files (full_path, filename, size_bytes)
        """
        if not identifiers:
            return []
        
        # Collect all matches, avoiding duplicates
        seen_paths = set()
        all_matches = []
        
        for identifier in identifiers:
            if not identifier:
                continue
            
            matches = self.match_identifier(str(identifier))
            for match in matches:
                full_path = match[0]
                if full_path not in seen_paths:
                    seen_paths.add(full_path)
                    all_matches.append(match)
        
        return all_matches
    
    def match_identifiers_and(
        self, 
        identifiers: List[str]
    ) -> List[Tuple[str, str, int]]:
        """
        Find files matching ALL of the given identifiers (AND logic).
        
        A file must contain ALL identifiers in its filename to match.
        
        Args:
            identifiers: List of identifiers - ALL must be present
            
        Returns:
            List of matched files (full_path, filename, size_bytes)
        """
        if not identifiers:
            return []
        
        # Filter out empty identifiers
        valid_identifiers = [str(id).strip() for id in identifiers if id and str(id).strip()]
        
        if not valid_identifiers:
            return []
        
        # Find files containing ALL identifiers (case-insensitive)
        matched = []
        valid_identifiers_lower = [id.lower() for id in valid_identifiers]
        for full_path, filename, size in self.file_index:
            filename_lower = filename.lower()
            # Check if ALL identifiers are present in filename
            all_present = all(identifier in filename_lower for identifier in valid_identifiers_lower)
            if all_present:
                matched.append((full_path, filename, size))
        
        return matched
    
    def get_match_statistics(
        self, 
        identifiers: List[str]
    ) -> Dict[str, Any]:
        """
        Get statistics about matching for a list of identifiers.
        
        Args:
            identifiers: List of identifiers to analyze
            
        Returns:
            Dictionary with statistics
        """
        total = len(identifiers)
        matched_count = 0
        unmatched = []
        total_matched_files = 0
        total_matched_size = 0
        
        for identifier in identifiers:
            if not identifier:
                continue
            
            matches = self.match_identifier(identifier)
            if matches:
                matched_count += 1
                total_matched_files += len(matches)
                total_matched_size += sum(size for _, _, size in matches)
            else:
                unmatched.append(identifier)
        
        return {
            'total_identifiers': total,
            'matched_count': matched_count,
            'unmatched_count': len(unmatched),
            'unmatched_identifiers': unmatched,
            'total_matched_files': total_matched_files,
            'total_matched_size': total_matched_size,
            'total_matched_size_formatted': format_bytes(total_matched_size),
            'match_percentage': (matched_count / total * 100) if total > 0 else 0,
        }
    
    def get_unmatched_files(
        self, 
        identifiers: List[str]
    ) -> List[Tuple[str, str, int]]:
        """
        Get files that don't match any identifier.
        
        Useful for identifying orphan files in the directory.
        
        Args:
            identifiers: List of identifiers that were used
            
        Returns:
            List of files not matched by any identifier
        """
        # Get all matched file paths
        matched_paths = set()
        for identifier in identifiers:
            if identifier:
                matches = self.match_identifier(identifier)
                for full_path, _, _ in matches:
                    matched_paths.add(full_path)
        
        # Find unmatched
        unmatched = [
            (full_path, filename, size)
            for full_path, filename, size in self.file_index
            if full_path not in matched_paths
        ]
        
        return unmatched
    
    def get_total_attachment_size(
        self, 
        identifier: str, 
        static_files: List[str] = None
    ) -> Tuple[int, bool, str]:
        """
        Calculate total attachment size for an identifier.
        
        Includes both matched files and static files.
        
        Args:
            identifier: Client identifier
            static_files: List of static attachment paths
            
        Returns:
            Tuple of (total_bytes, exceeds_limit, formatted_size)
        """
        total = 0
        
        # Add matched files
        matches = self.match_identifier(identifier)
        for _, _, size in matches:
            total += size
        
        # Add static files
        if static_files:
            for file_path in static_files:
                try:
                    total += os.path.getsize(file_path)
                except OSError:
                    pass
        
        exceeds = total > config.MAX_ATTACHMENT_SIZE_BYTES
        formatted = format_bytes(total)
        
        return total, exceeds, formatted
    
    def generate_match_report(
        self, 
        identifiers: List[str],
        output_path: str = None
    ) -> Dict[str, Any]:
        """
        Generate a detailed match report.
        
        Args:
            identifiers: List of identifiers
            output_path: Optional path to save Excel report
            
        Returns:
            Dictionary containing report data
        """
        report = {
            'generated_at': datetime.now().isoformat(),
            'directory': self.directory,
            'recursive': self.recursive,
            'total_files_scanned': self.total_files,
            'total_size_scanned': format_bytes(self.total_size),
            'identifiers': [],
            'summary': None,
        }
        
        # Process each identifier
        for identifier in identifiers:
            if not identifier:
                continue
            
            matches = self.match_identifier(identifier)
            total_size = sum(size for _, _, size in matches)
            
            report['identifiers'].append({
                'identifier': str(identifier),
                'match_count': len(matches),
                'total_size': format_bytes(total_size),
                'files': [
                    {
                        'filename': filename,
                        'size': format_bytes(size),
                        'path': full_path
                    }
                    for full_path, filename, size in matches
                ]
            })
        
        # Add summary
        stats = self.get_match_statistics(identifiers)
        report['summary'] = stats
        
        # Optionally save to Excel
        if output_path:
            self._save_report_to_excel(report, output_path)
        
        return report
    
    def _save_report_to_excel(self, report: Dict, output_path: str) -> bool:
        """
        Save match report to Excel file.
        
        Args:
            report: Report dictionary
            output_path: Path for Excel file
            
        Returns:
            True if successful
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            ws_summary['A1'] = "Match Report Summary"
            ws_summary['A1'].font = Font(bold=True, size=14)
            
            summary_data = [
                ("Generated At", report['generated_at']),
                ("Directory", report['directory']),
                ("Recursive Search", str(report['recursive'])),
                ("Total Files Scanned", report['total_files_scanned']),
                ("Total Size Scanned", report['total_size_scanned']),
                ("", ""),
                ("Identifiers Processed", report['summary']['total_identifiers']),
                ("Identifiers with Matches", report['summary']['matched_count']),
                ("Identifiers without Matches", report['summary']['unmatched_count']),
                ("Match Percentage", f"{report['summary']['match_percentage']:.1f}%"),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, start=3):
                ws_summary.cell(row=row_idx, column=1, value=label)
                ws_summary.cell(row=row_idx, column=2, value=value)
            
            # Details sheet
            ws_details = wb.create_sheet("Match Details")
            headers = ["Identifier", "File Count", "Total Size", "Filenames"]
            for col, header in enumerate(headers, start=1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, id_data in enumerate(report['identifiers'], start=2):
                ws_details.cell(row=row_idx, column=1, value=id_data['identifier'])
                ws_details.cell(row=row_idx, column=2, value=id_data['match_count'])
                ws_details.cell(row=row_idx, column=3, value=id_data['total_size'])
                filenames = "; ".join(f['filename'] for f in id_data['files'])
                ws_details.cell(row=row_idx, column=4, value=filenames)
            
            # Unmatched sheet
            if report['summary']['unmatched_identifiers']:
                ws_unmatched = wb.create_sheet("Unmatched")
                ws_unmatched.cell(row=1, column=1, value="Unmatched Identifiers").font = Font(bold=True)
                for row_idx, identifier in enumerate(report['summary']['unmatched_identifiers'], start=2):
                    ws_unmatched.cell(row=row_idx, column=1, value=identifier)
            
            wb.save(output_path)
            self.logger.info(f"Match report saved to: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving report: {e}", exc_info=True)
            return False
    
    def clear(self) -> None:
        """Clear all cached data."""
        self.directory = None
        self.recursive = False
        self.file_index = []
        self.matches = {}
        self.scan_time = None
        self.total_files = 0
        self.total_size = 0
